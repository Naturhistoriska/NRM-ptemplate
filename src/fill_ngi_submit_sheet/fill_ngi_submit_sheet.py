#! /usr/bin/env python3
# vim:fenc=utf-8
#
# Copyright © 2024 nylander <johan.nylander@nrm.se>
#
# Distributed under terms of the MIT license.

"""
Fill NGI sequencing submit sheet from ENA_fastq2_template.tsv and Illumina_unique_dual_index_seqs.xlsx

Input: NGI_submit_sheet.xlsx

Output: NGI_submit_sheet.filled.xlsx
"""

import os
import sys
import argparse
import shutil
import openpyxl
import pandas as pd

openpyxl.reader.excel.warnings.simplefilter(action='ignore')

ENA_URL = "https://nrmcloud.nrm.se/s/fqTaa7ezsgKjmcG/download"      # xlsx
ILLUMINA_URL = "https://nrmcloud.nrm.se/s/T3GgxSJyKeyxSwo/download" # xlsx
ILLUMINA_SHEET_NAME = "Illumina info"
NGI_SHEET_NAME = "Sample information"
WELL_DEF = "A1"
SAMPLE_TYPE_DEF = "Finished library" # "Finished library", "Amplicon with adapters (low diversity)"
SEQUENCING_PLATFORM_DEF = "NovaSeq"  # NovaSeq, MiSeq, NextSeq

VERSION = '0.1'

def get_options():
    """
    Get options from command line
    """
    parser = argparse.ArgumentParser(
            description='fill ngi submit sheet',
            formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('ngi_file', help='Input NGI excel file.')
    parser.add_argument(
            '-e', '--ena-file', dest='ena_file', default=ENA_URL,
            help='ENA sample sheet.')
    parser.add_argument(
            '-i', '--illumina-file', dest='illumina_file', default=ILLUMINA_URL,
           help='Illumina barcode sample sheet.')
    parser.add_argument(
            '-w', '--well', dest='well', default=WELL_DEF,
            help='well.')
    parser.add_argument(
            '-t', '--sample-type', dest='sample_type',
            choices=['Finished library', 'Amplicon with adapters (low diversity)'],
            default=SAMPLE_TYPE_DEF,
            help='sample type.')
    parser.add_argument(
            '-p', '--sequencing-platform', dest='sequencing_platform',
            choices=['NovaSeq', 'MiSeq', 'NextSeq'],
            default=SEQUENCING_PLATFORM_DEF,
            help='Sequencing platform.')
    parser.add_argument(
            '-v', '--version', dest='version', action='version', version=VERSION,
            help='print version and exit.')
    options = parser.parse_args()
    return options


if __name__ == '__main__':

    opts = get_options()

    if opts.ngi_file:
        print(f"Reading {opts.ngi_file}")
        ngi_base_name, ngi_ext = os.path.splitext(opts.ngi_file)
        fill_file = ngi_base_name + '.filled' + ngi_ext
        if os.path.exists(fill_file):
            print(f"Warning: A filled copy of {opts.ngi_file} seems to exists ({fill_file}). Will not overwrite.")
            sys.exit(0)
        else:
            shutil.copy2(opts.ngi_file, fill_file)

    if opts.ena_file:
        print(f"Reading {opts.ena_file}")
        if os.path.exists(opts.ena_file):
            ena_base_name, ena_ext = os.path.splitext(opts.ena_file)
            if ena_ext.lower() == '.xlsx':
                ENA = pd.read_excel(opts.ena_file)
            elif ena_ext.lower() == '.tsv':
                ENA = pd.read_csv(opts.ena_file, sep='\t')
        else:
            ENA = pd.read_excel(opts.ena_file) # assuming URL is xlsx

    n_samples = len(ENA) - 3

    if opts.illumina_file:
        print(f"Reading {opts.illumina_file}")
        ILL = pd.read_excel(opts.illumina_file, sheet_name=ILLUMINA_SHEET_NAME)

    # Local excel file and sheet
    wb = openpyxl.load_workbook(fill_file)
    ws = wb[NGI_SHEET_NAME]

    # From ENA, read all values from column 3, starting at row 4.
    # Insert values in local file NGI.filled.xlsx,
    # sheet "Sample information", column Q, starting at cell Q20.
    for i, val in enumerate(ENA.iloc[3:, 2]):
        ws[f"Q{20+i}"] = val

    # From ILL, read all values from column A, starting at row 2.
    # Insert values in local file NGI.filled.xlsx,
    # sheet "Sample information", column S, starting at cell S20.
    for i, val in enumerate(ILL.iloc[1:, 0]):
        ws[f"S{20+i}"] = val
        if i == n_samples-1:
            break

    # From ILL, read all values from column H, starting at row 2.
    # Insert values in local file NGI.filled.xlsx,
    # sheet "Sample information", column U, starting at cell U20.
    for i, val in enumerate(ILL.iloc[1:, 7]):
        ws[f"U{20+i}"] = val
        if i == n_samples-1:
            break

    # Insert string "Custom" in local file NGI.filled.xlsx,
    # sheet "Sample information", column R, starting at cell R20.
    for i in range(n_samples):
        ws[f"R{20+i}"] = "Custom"

    # Insert string "A1" in local file NGI.filled.xlsx,
    # sheet "Sample information", column P, starting at cell P20.
    for i in range(n_samples):
        ws[f"P{20+i}"] = opts.well

    # Insert string "Finished library" in local file NGI.filled.xlsx,
    # sheet "Sample information", cell Q8
    ws["Q8"] = opts.sample_type

    # Insert string "NovaSeq" in local file NGI.xlsx,
    # sheet "Sample information", cell S8
    ws["S8"] = opts.sequencing_platform

    # Save
    wb.save(fill_file)
    print(f"Wrote {fill_file}\nEnd of script.")
